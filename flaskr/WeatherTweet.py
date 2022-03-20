#note: time info used in variable name is JST based(e.g.: post_content_monday_2113)
#for executing on web server 
import cgi
import sys, io
import os
import tweepy
#coding: utf-8
import requests
import json
import time
import datetime
import schedule
import functools
import random
import emoji
#for db connection
import sqlite3
#for importing Excel data
from openpyxl import load_workbook
from contextlib import closing
import shutil

PF = os.getenv("PROJECT_FOLDER")
sys.path.append(PF)

##Tweet class
##keys & tokens for authorization
class Tweet():
# Insert keys from Twitter account
    CK = os.getenv("CONSUMER_KEY")
    CS = os.getenv("CONSUMER_SECRET")
    AT = os.getenv("ACCESS_TOKEN")
    AS = os.getenv("ACCESS_TOKEN_SECRET")
    #generate Twitter object
    auth = tweepy.OAuthHandler(CK, CS)
    auth.set_access_token(AT, AS)
    api = tweepy.API(auth)
    recent_posts_tmp = api.user_timeline()[:5]
    recent_posts=[]
    for i in range(5):
        recent_posts.append(recent_posts_tmp[i]._json["text"])
    #init
    def __init__(self,post):
        pass
    #Tweet
    def Tweet(self,post):
        self.api.update_status(emoji.emojize(post))
    #Remove the same tweets.
    def RmDupTweet(self,content):
        possible_posts = content
        remained_posts_tmp = [i for i in possible_posts if i not in self.recent_posts]
        remained_posts = [a for a in remained_posts_tmp if a != '']
        return remained_posts

#WeatherTweet class
class WeatherTweet():
    count =0
    ##tweet time will be chosen randomly.##
    ##Tweet time (hours, UTC based)##
    hours = [1,2,3,4,5,6,7,8,9,10,11,12,13]
    ##Tweet time (minutes)##
    minutes = list(range(60))

    ##import data from Excel file "Shibu_Tweet.xlsx" if it exists in the same directory
    DB_PATH=os.getenv("DB_PATH")
    PATH = os.path.join(os.getenv("UPLOAD_FOLDER"),'Shibu_Tweet.xlsx')
    if os.path.isfile(PATH) == True:
        infile = PATH
        wb = load_workbook(filename = infile)
        ws = wb['Sheet1']
        r_lst=[]
        for row in range(3, ws.max_row + 1):
            r_tpl = ()
            for col in range(1, 20):
                if ws.cell(row,col).value == None:
                    ws.cell(row,col).value = ""
                else:
                    pass
                r_tpl += (ws.cell(row,col).value, )
            r_lst.append(r_tpl)

        ##conversion from jst to utc
        #extract first 2digit of r_lst(jst_h)
        jst = [i[4] for i in r_lst]
        jst_h = [j[:j.find(":")] for j in jst]
        utc_m = [k[-2:] for k in jst]
        utc_h = [ str(int(l)+15) if int(l) <= 8 else str(int(l)-9).zfill(2) for l in jst_h]
        utc = []
        for m in range(0,len(r_lst)):
            utc.append(utc_h[m] + ":" + utc_m[m])
        print("utc=",utc)
        ##end conversion jst to utc
        #replace tw_time date with utc
        for o in range(0,len(r_lst)):
            tmp1 = list(r_lst[o])
            tmp1[4] = utc[o]
            r_lst[o] = tuple(tmp1)
        for row in range(0, ws.max_row - 2):
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            uid = c.execute(
                ' SELECT id '
                ' FROM user '
                ' WHERE username = ?',
                (r_lst[row][18],)
                ).fetchone()
            tmp2 = list(r_lst[row])
            tmp2[18] = uid[0]
            r_lst[row] = tuple(tmp2)
            c.execute(
                ' INSERT INTO tweet_table (tw_type, tw_memo, tw_dow, tw_rdm, tw_time, weather_condition, temp_condition, tweet1, tweet2, tweet3, tweet4, tweet5, tweet6, tweet7, tweet8, tweet9, tweet10, created_at, author_id)'
                ' VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) ',
                (r_lst[row])
                )
            conn.commit()
            conn.close()
    else:
        pass
    #clear uploads folder after reading
    target_dir = os.getenv("UPLOAD_FOLDER")
    shutil.rmtree(target_dir)
    os.mkdir(target_dir)

    ##fetch db##
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    data = c.execute(
        'SELECT tt.id, tw_type, tw_memo, tw_dow, tw_rdm, tw_time, weather_condition, temp_condition, tweet1, tweet2,tweet3,tweet4,tweet5,tweet6,tweet7,tweet8,tweet9,tweet10, created_at, author_id, username'
        ' FROM tweet_table tt JOIN user u ON tt.author_id = u.id'
        ' ORDER BY tw_dow ASC, tw_time ASC'
        ).fetchall()
    conn.close()

    #for test tweet
    test = ["This is test tweet1","This is test tweet2","This is test tweet3","This is test tweet4","This is test tweet5"]

    def __init__(self):
        self.current_time = ""
        self.current_weather = ""
        self.current_temp = 0
        self.tweet = Tweet("")

    #A decorator that catches errors of schedule function
    def catch_exceptions(cancel_on_failure=False):
        def catch_exceptions_decorator(job_func):
            @functools.wraps(job_func)
            def wrapper(*args, **kwargs):
                try:
                    return job_func(*args, **kwargs)
                except:
                    import traceback
                    print(traceback.format_exc())
                    if cancel_on_failure:
                        return schedule.CancelJob
            return wrapper
        return catch_exceptions_decorator

    #Tweet start time (Daily tweet at fixed time.)
    @catch_exceptions(cancel_on_failure=True)
    def main(self):
        i=0
        for row in self.data:
            i = i+1
            tw_dow=row[3]
            tw_rdm=row[4]
            tw_time=row[5]
            #tw_hour="{:.2}".format(tw_time)
            tw_hour=tw_time[:2]
            tw_minute=tw_time[3:]
            tw_post=[]
            for i in range(8,18):
                tw_post.append(row[i])

            if row[1]=="Daily":
                #daily post(random_hour)
                if tw_rdm=="random_hour":
                    schedule.every().day.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.daily_post,tw_post)
                #daily post(random_minute) 
                elif tw_dow=="random_minute":
                    schedule.every().day.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.daily_post,tw_post)
                #daily post(random_hour_minute)      
                elif tw_rdm=="random_hour_minute":
                    schedule.every().day.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.daily_post,tw_post)
                #daily post(normal)
                elif tw_rdm=="not_random":
                    schedule.every().day.at(str(tw_time)).do(self.daily_post,tw_post)

            elif row[1]=="Weekly":
                #print("day=",tw_dow)
                #weekly post(random_hour)
                if tw_rdm=="random_hour":
                    self.weekly_post_schedule_random_hour(tw_dow,tw_minute,tw_post)
                #weekly post(random_minute)
                elif tw_rdm=="random_minute":
                    self.weekly_post_schedule_random_minute(tw_dow,tw_hour,tw_post)
                    #schedule.every().tw_dow.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weekly_post,tw_post)
                elif tw_rdm=="random_hour_minute":
                    self.weekly_post_schedule_random_hour_minute(tw_dow,tw_post)
                #weekly post(normal)
                elif tw_rdm=="not_random":
                    self.weekly_post_schedule(tw_dow,tw_time,tw_post)
                    #schedule.every().tw_dow.at(tw_time).do(self.weekly_post,tw_post)
            else:
                weather_cond=row[6]
                temp_cond=row[7]
                #weather_post(random_hour)
                if tw_rdm=="random_hour":
                    self.weather_post_schedule_random_hour(tw_dow,tw_minute,weather_cond,temp_cond,tw_post)
                #weather_post(random_minute)
                elif tw_rdm=="random_minute":
                    self.weather_post_schedule_random_minute(tw_dow,tw_hour,weather_cond,temp_cond,tw_post)
                    #schedule.every().tw_dow.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post)
                elif tw_rdm=="random_hour_minute":
                    self.weather_post_schedule_random_hour_minute(tw_dow,weather_cond,temp_cond,tw_post)
                #weather_post(normal)
                elif tw_rdm=="not_random":
                    self.weather_post_schedule(tw_dow,tw_time,weather_cond,temp_cond,tw_post)

        #post 1 question per day among "question" list at random timing
        #schedule.every().day.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,self.question)

        #for test (JST base)
        #schedule.every().day.at("18:"+str(f'{random.choice(self.minutes):02}')).do(self.post_test)
        #schedule.every().day.at("16:56").do(self.post_test)
        while True:
            schedule.run_pending()
            time.sleep(1)
            self.count = self.count+1
            print(self.count, "秒")

    #Stop tweet
    def stop(self):
        sys.exit("stopped auto tweet")

    #Template of post function
    def post_base(self,post):
        remained_posts=self.tweet.RmDupTweet(post)
        #print(random.choice(remained_posts))
        self.tweet.Tweet(random.choice(remained_posts))

    #Daily post
    def daily_post(self,post):
        self.post_base(post)

    #weekly post schedule(post hour is random)
    def weekly_post_schedule_random_hour(self,tw_dow,tw_minute,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)
        else:
            schedule.every().sunday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.post_base,tw_post)

    #weekly post schedule(post minute is random)
    def weekly_post_schedule_random_minute(self,tw_dow,tw_hour,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        else:
            schedule.every().sunday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)

    #weekly post schedule(post hour&minute is random)
    def weekly_post_schedule_random_hour_minute(self,tw_dow,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)
        else:
            schedule.every().sunday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.post_base,tw_post)

    #weekly post schedule
    def weekly_post_schedule(self,tw_dow,tw_time,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(tw_time).do(self.post_base,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(tw_time).do(self.post_base,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(tw_time).do(self.post_base,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(tw_time).do(self.post_base,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(tw_time).do(self.post_base,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(tw_time).do(self.post_base,tw_post)
        else:
            schedule.every().sunday.at(tw_time).do(self.post_base,tw_post)

    #Weather related post
    def weather_post(self,weather_cond,temp_cond,tw_post):
        self.get_weather_info()
        if self.weather_check(weather_cond) == True:
            if self.temp_check(temp_cond) == True:
                self.post_base(tw_post)
            else:
                pass
        else:
            pass
    #weather post schedule(post hour is random)
    def weather_post_schedule_random_hour(self,tw_dow,tw_minute,weather_cond,temp_cond,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
        else:
            schedule.every().sunday.at(str(f'{random.choice(self.hours):02}')+":"+str(tw_minute)).do(self.weather_post,weather_cond,temp_cond,tw_post)
    #weather post schedule(post minute is random)
    def weather_post_schedule_random_minute(self,tw_dow,tw_hour,weather_cond,temp_cond,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        else:
            schedule.every().sunday.at(str(tw_hour)+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
    def weather_post_schedule_random_hour_minute(self,tw_dow,weather_cond,temp_cond,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
        else:
            schedule.every().sunday.at(str(f'{random.choice(self.hours):02}')+":"+str(f'{random.choice(self.minutes):02}')).do(self.weather_post,weather_cond,temp_cond,tw_post)
    #weather post schedule
    def weather_post_schedule(self,tw_dow,tw_time,weather_cond,temp_cond,tw_post):
        if tw_dow=="Monday":
            schedule.every().monday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Tuesday":
            schedule.every().tuesday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Wednesday":
            schedule.every().wednesday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Thursday":
            schedule.every().thursday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Friday":
            schedule.every().friday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
        elif tw_dow=="Saturday":
            schedule.every().saturday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
        else:
            schedule.every().sunday.at(tw_time).do(self.weather_post,weather_cond,temp_cond,tw_post)
    #for test tweet
    def post_test(self):
        self.post_base(self.test)

    #Get weather data from OpenWeatherMap
    def get_weather_info(self):
        city_name = os.getenv("CITY_NAME")
        API_KEY = os.getenv("API_KEY")
        api = "http://api.openweathermap.org/data/2.5/weather?units=metric&q={city}&APPID={key}"
        url = api.format(city = city_name, key = API_KEY)
        response = requests.get(url)
        data = response.json() #dict. type
        #store obtained weather&temperature data
        self.current_time = datetime.datetime.now()
        self.current_weather = data["weather"][0]["main"]
        self.current_temp = float(data["main"]["temp"])
        print("current_time=",self.current_time)
        print("current_weather=",self.current_weather)
        print("current_temp=",self.current_temp)
    
    #select post content according to weather info
    #https://openweathermap.org/weather-conditions
    def weather_check(self,weather_cond):
        if weather_cond == "Clear":
            if self.current_weather == "Clear":
                return True
            else:
                return False
        elif weather_cond == "Clouds":
            if self.current_weather == "Clouds":
                return True
            else:
                return False
        elif weather_cond == "Rainy":
            if self.current_weather == "Thunderstorm" or self.current_weather == "Drizzle" or self.current_weather == "Rain":
                return True
            else:
                return False
        elif weather_cond == "Others":
            if self.current_weather in ["Snow", "Mist", "Smoke","Haze","Dust","Fog","Sand","Ash", "Squall", "Tornado"]    :
                return True
            else:
                return False
        #elif weather_cond == "NoWeatherCondition"
        else:
            return True

    def temp_check(self,temp_cond):
        if temp_cond == "below5":
            if self.current_temp <= 5:
                return True
            else:
                return False
        elif temp_cond == "5-10":
            if self.current_temp > 5 and self.current_temp <=10:
                return True
            else:
                return False
        elif temp_cond == "10-15":
            if self.current_temp > 10 and self.current_temp <=15:
                return True
            else:
                return False
        elif temp_cond == "15-20":
            if self.current_temp > 15 and self.current_temp <=20:
                return True
            else:
                return False
        elif temp_cond == "20-25":
            if self.current_temp > 20 and self.current_temp <=25:
                return True
            else:
                return False
        elif temp_cond == "25-30":
            if self.current_temp > 25 and self.current_temp <=30:
                return True
            else:
                return False
         #elif temp_cond == "over30":
        else:
            if self.current_temp > 30:
                return True
            else:
                return False
#for test
if __name__ == "__main__":
    w = WeatherTweet()
    w.main()